import os
import re
import csv
import zipfile
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
import xml.etree.ElementTree as ET
import warnings
warnings.filterwarnings('ignore')

from speed_recognizer import SpeedRecognizer


class DataProcessor:
    def __init__(self, task_dir: Path):
        self.task_dir = task_dir
        self.recognizer = SpeedRecognizer()
        self.log_4g_dir = None
        self.log_5g_dir = None
        
        # 查找log目录
        for subdir in task_dir.rglob('*'):
            if subdir.is_dir():
                if '4G测试log' in subdir.name and 'cellular' in subdir.name:
                    self.log_4g_dir = subdir
                elif '5G测试log' in subdir.name and 'cellular' in subdir.name:
                    self.log_5g_dir = subdir
    
    def extract_images_from_excel(self, excel_path: Path) -> Dict[str, Path]:
        """从Excel中提取图片，返回{descr: img_path}和{dispimg_id: img_path}的映射"""
        images = {}
        self.dispimg_to_image = {}  # DISPIMG ID -> 图片路径
        
        try:
            img_dir = self.task_dir / 'extracted_images'
            img_dir.mkdir(exist_ok=True)
            
            with zipfile.ZipFile(excel_path, 'r') as zip_ref:
                # 1. 提取所有图片到临时目录，建立文件名到路径的映射
                image_files_by_name = {}  # 如 'image93.jpeg' -> Path
                for file_info in zip_ref.filelist:
                    if 'xl/media/' in file_info.filename:
                        if file_info.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                            img_data = zip_ref.read(file_info.filename)
                            img_name = Path(file_info.filename).name
                            img_path = img_dir / img_name
                            with open(img_path, 'wb') as f:
                                f.write(img_data)
                            image_files_by_name[img_name] = img_path
                
                print(f"提取了 {len(image_files_by_name)} 个图片文件")
                
                # 2. 解析 cellimages.xml.rels 获取 rId -> 图片文件名 的映射
                rels_xml = None
                rid_to_image = {}  # 如 'rId93' -> 'image93.jpeg'
                try:
                    rels_xml = zip_ref.read('xl/_rels/cellimages.xml.rels')
                except:
                    pass
                
                if rels_xml:
                    try:
                        rels_root = ET.fromstring(rels_xml)
                        rels_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                        for rel in rels_root.findall('.//r:Relationship', rels_ns):
                            rid = rel.get('Id')
                            target = rel.get('Target', '')
                            if rid and 'media/' in target:
                                img_name = Path(target).name
                                rid_to_image[rid] = img_name
                        # 如果命名空间不对，尝试不带命名空间
                        if not rid_to_image:
                            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                rid = rel.get('Id')
                                target = rel.get('Target', '')
                                if rid and 'media/' in target:
                                    img_name = Path(target).name
                                    rid_to_image[rid] = img_name
                        print(f"解析rels文件，找到 {len(rid_to_image)} 个rId映射")
                    except Exception as e:
                        print(f"解析cellimages.xml.rels失败: {e}")
                
                # 3. 解析 cellimages.xml 获取 name(DISPIMG ID) -> rId 和 descr -> rId 的映射
                cellimages_xml = None
                try:
                    cellimages_xml = zip_ref.read('xl/cellimages.xml')
                except:
                    try:
                        cellimages_xml = zip_ref.read('xl/docs/cellimages.xml')
                    except:
                        pass
                
                if cellimages_xml:
                    try:
                        root = ET.fromstring(cellimages_xml)
                        ns = {
                            'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                        }
                        
                        cell_images = root.findall('.//etc:cellImage', ns)
                        print(f"XML中找到 {len(cell_images)} 个cellImage节点")
                        
                        for cell_image in cell_images:
                            pic = cell_image.find('.//xdr:pic', ns)
                            if pic is None:
                                continue
                            
                            c_nv_pr = pic.find('.//xdr:cNvPr', ns)
                            blip = pic.find('.//a:blip', ns)
                            
                            if c_nv_pr is None or blip is None:
                                continue
                            
                            name = c_nv_pr.get('name', '')  # DISPIMG ID
                            descr = c_nv_pr.get('descr', '')  # 描述
                            embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '')
                            
                            if not embed:
                                continue
                            
                            # 通过 rId 获取图片文件名，再获取图片路径
                            img_name = rid_to_image.get(embed)
                            if img_name and img_name in image_files_by_name:
                                img_path = image_files_by_name[img_name]
                                
                                # 建立 descr -> 图片 的映射
                                if descr:
                                    images[descr] = img_path
                                
                                # 建立 DISPIMG ID (name) -> 图片 的映射
                                if name:
                                    self.dispimg_to_image[name] = img_path
                        
                        print(f"成功建立 {len(images)} 个descr映射，{len(self.dispimg_to_image)} 个DISPIMG ID映射")
                    except Exception as e:
                        print(f"解析cellimages.xml失败: {e}")
                        import traceback
                        traceback.print_exc()
                else:
                    print("未找到cellimages.xml文件")
            
        except Exception as e:
            print(f"提取图片失败: {e}")
        
        return images
    
    def get_image_by_dispimg_id(self, dispimg_id: str) -> Optional[Path]:
        """通过DISPIMG ID获取图片路径"""
        if not hasattr(self, 'dispimg_to_image'):
            return None
        return self.dispimg_to_image.get(dispimg_id)
    
    def get_image_for_row(self, row_data: Dict, column_name: str, images: Dict) -> Optional[Path]:
        """根据行数据获取对应的图片，严格匹配，找不到返回None"""
        if not images:
            return None
        
        # 尝试通过log文件名匹配
        log_4g_1 = row_data.get('4G测试log（cellular）_1', '')
        log_4g_2 = row_data.get('4G测试log（cellular）_2', '')
        log_5g_1 = row_data.get('5G测试log（cellular）_1', '')
        log_5g_2 = row_data.get('5G测试log（cellular）_2', '')
        
        # 提取log文件名（去掉扩展名，将冒号替换为下划线）
        log_names = []
        for log in [log_4g_1, log_4g_2, log_5g_1, log_5g_2]:
            if log:
                log_name = str(log).strip()
                if log_name:
                    log_name = log_name.replace(':', '_').replace('.csv', '').replace('.CSV', '')
                    log_names.append(log_name)
        
        # 确定目标关键词
        target_keywords = []
        if '5G速率图' in column_name:
            target_keywords = ['5G速率图']
        elif '4G速率图' in column_name:
            target_keywords = ['4G速率图']
        
        # 严格匹配：必须同时满足时间戳+手机号匹配 和 关键词匹配
        for log_name in log_names:
            # 提取时间戳和手机号：2025_11_22 11:08_13392507898
            parts = log_name.split('--')
            if len(parts) > 0:
                time_phone = parts[0].strip()
                if not time_phone:
                    continue
                
                # 在images中查找同时匹配时间戳+手机号和关键词的图片
                for descr, img_path in images.items():
                    # 必须包含时间戳+手机号
                    if time_phone not in descr:
                        continue
                    # 必须包含目标关键词
                    if not any(kw in descr for kw in target_keywords):
                        continue
                    # 严格匹配成功
                    return img_path
        
        # 找不到匹配的图片，返回None
        return None
    
    def process_excel(self, excel_path: Path, progress_callback=None) -> List[Dict]:
        """处理Excel文件，提取数据
        
        Args:
            excel_path: Excel文件路径
            progress_callback: 进度回调函数，签名为 callback(current, total, result)
                              current: 当前处理的行号
                              total: 总行数
                              result: 当前行的处理结果
        """
        results = []
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # 读取表头
        headers = []
        for col_idx, cell in enumerate(ws[1], 1):
            headers.append(str(cell.value) if cell.value else f'Col{col_idx}')
        
        # 查找关键列的索引
        col_indices = {}
        target_cols = [
            '工单号',
            '5G速率图（移动爱家）',
            # 新字段名称（合并后）
            '4G测试log（cellular）',
            '5G测试log（cellular）',
            # 旧字段名称（向后兼容）
            '4G测试log（cellular）_1',
            '4G测试log（cellular）_2',
            '5G测试log（cellular）_1',
            '5G测试log（cellular）_2'
        ]
        
        for col_name in target_cols:
            # 先尝试精确匹配
            for idx, header in enumerate(headers):
                if header.strip() == col_name:
                    col_indices[col_name] = idx + 1
                    break
            # 如果没找到，尝试模糊匹配（列名包含目标字符串）
            if col_name not in col_indices:
                for idx, header in enumerate(headers):
                    if col_name in header:
                        col_indices[col_name] = idx + 1
                        break
        
        if '工单号' not in col_indices:
            raise Exception("未找到'工单号'列")
        
        # 提取图片
        images = self.extract_images_from_excel(excel_path)
        
        # 先统计总行数
        all_rows = list(ws.iter_rows(min_row=2))
        valid_rows = [row for row in all_rows if row[col_indices['工单号'] - 1].value]
        total_rows = len(valid_rows)
        
        # 处理每一行数据
        for idx, row in enumerate(valid_rows):
            order_id = str(row[col_indices['工单号'] - 1].value)
            result = {'工单号': order_id}
            
            # 构建行数据字典
            row_data = {}
            for col_name in col_indices:
                cell = row[col_indices[col_name] - 1]
                cell_value = cell.value
                row_data[col_name] = str(cell_value).strip() if cell_value else ''
            
            # 处理5G速率图 - 支持从DISPIMG公式中提取图片
            speed_5g_img = None
            if '5G速率图（移动爱家）' in col_indices:
                speed_5g_cell_value = row_data.get('5G速率图（移动爱家）', '')
                
                # 检查是否是DISPIMG公式
                if 'DISPIMG' in speed_5g_cell_value.upper():
                    # 提取DISPIMG ID: =DISPIMG("ID_xxx",1) 或 =_xlfn.DISPIMG("ID_xxx",1)
                    match = re.search(r'DISPIMG\s*\(\s*"([^"]+)"', speed_5g_cell_value, re.IGNORECASE)
                    if match:
                        dispimg_id = match.group(1)
                        speed_5g_img = self.get_image_by_dispimg_id(dispimg_id)
                        if speed_5g_img:
                            print(f"  通过DISPIMG ID找到图片: {dispimg_id} -> {speed_5g_img.name}")
                
                # 如果DISPIMG方式没找到，尝试通过descr匹配
                if not speed_5g_img:
                    speed_5g_img = self.get_image_for_row(row_data, '5G速率图（移动爱家）', images)
            
            if speed_5g_img and speed_5g_img.exists():
                speeds = self.recognizer.recognize_image(str(speed_5g_img))
                result['上传速率Mbps'] = speeds.get('upload_speed')
                result['下载速率Mbps'] = speeds.get('download_speed')
            else:
                result['上传速率Mbps'] = None
                result['下载速率Mbps'] = None
            
            # 处理4G log - 优先使用新字段名称，兼容旧字段名称
            log_4g_name = None
            # 新字段名称（合并后）
            log_4g = row_data.get('4G测试log（cellular）', '').strip()
            if log_4g and not (log_4g.startswith('=_') or log_4g.startswith('=')):
                log_4g_name = log_4g
            else:
                # 旧字段名称（向后兼容）
                log_4g_1 = row_data.get('4G测试log（cellular）_1', '').strip()
                log_4g_2 = row_data.get('4G测试log（cellular）_2', '').strip()
                if log_4g_1 and not (log_4g_1.startswith('=_') or log_4g_1.startswith('=')):
                    log_4g_name = log_4g_1
                elif log_4g_2 and not (log_4g_2.startswith('=_') or log_4g_2.startswith('=')):
                    log_4g_name = log_4g_2
            
            # 将文件名中的冒号替换为下划线（Excel中可能是11:08，实际文件是11_08）
            if log_4g_name:
                log_4g_name = log_4g_name.replace(':', '_')
            
            # 解析4G log数据
            log_4g_data = None
            if log_4g_name and self.log_4g_dir:
                log_4g_data = self.parse_4g_log(log_4g_name)
                result['ECI'] = log_4g_data.get('ECI')
                result['RSRP'] = log_4g_data.get('RSRP')
                result['SINR'] = log_4g_data.get('SINR')
            else:
                result['ECI'] = None
                result['RSRP'] = None
                result['SINR'] = None
            
            # 处理5G log - 优先使用新字段名称，兼容旧字段名称
            log_5g_name = None
            # 新字段名称（合并后）
            log_5g = row_data.get('5G测试log（cellular）', '').strip()
            if log_5g and not (log_5g.startswith('=_') or log_5g.startswith('=')):
                log_5g_name = log_5g
            else:
                # 旧字段名称（向后兼容）
                log_5g_1 = row_data.get('5G测试log（cellular）_1', '').strip()
                log_5g_2 = row_data.get('5G测试log（cellular）_2', '').strip()
                if log_5g_1 and not (log_5g_1.startswith('=_') or log_5g_1.startswith('=')):
                    log_5g_name = log_5g_1
                elif log_5g_2 and not (log_5g_2.startswith('=_') or log_5g_2.startswith('=')):
                    log_5g_name = log_5g_2
            
            # 将文件名中的冒号替换为下划线（Excel中可能是11:08，实际文件是11_08）
            if log_5g_name:
                log_5g_name = log_5g_name.replace(':', '_')
            
            # 解析5G log数据
            log_5g_data = None
            if log_5g_name and self.log_5g_dir:
                log_5g_data = self.parse_5g_log(log_5g_name)
                result['NR-CI'] = log_5g_data.get('NR-CI')
                result['SS-RSRP'] = log_5g_data.get('SS-RSRP')
                result['SS-SINR'] = log_5g_data.get('SS-SINR')
            else:
                result['NR-CI'] = None
                result['SS-RSRP'] = None
                result['SS-SINR'] = None
            
            # 经纬度处理：优先使用5G Log的经纬度，没有才用4G的
            result['经度'] = None
            result['纬度'] = None
            
            # 先检查5G Log是否有经纬度
            if log_5g_data:
                lon_5g = log_5g_data.get('5G_经度')
                lat_5g = log_5g_data.get('5G_纬度')
                if lon_5g and lat_5g:
                    result['经度'] = lon_5g
                    result['纬度'] = lat_5g
                    print(f"  使用5G经纬度: {lon_5g}, {lat_5g}")
            
            # 如果5G没有经纬度，使用4G的
            if not result['经度'] or not result['纬度']:
                if log_4g_data:
                    lon_4g = log_4g_data.get('经度')
                    lat_4g = log_4g_data.get('纬度')
                    if lon_4g and lat_4g:
                        result['经度'] = lon_4g
                        result['纬度'] = lat_4g
                        print(f"  使用4G经纬度: {lon_4g}, {lat_4g}")
            
            results.append(result)
            
            # 调用进度回调，如果返回 False 则停止处理
            if progress_callback:
                should_continue = progress_callback(idx + 1, total_rows, result)
                if should_continue is False:
                    print(f"处理被取消，已处理 {idx + 1}/{total_rows} 行")
                    break
        
        return results
    
    def parse_4g_log(self, log_filename: str) -> Dict:
        """解析4G log文件（支持CSV和XLSX格式）
        优先选择有经纬度的数据，没有才选择无经纬度的数据
        """
        result = {
            '经度': None,
            '纬度': None,
            'ECI': None,
            'RSRP': None,
            'SINR': None
        }
        
        print(f"处理4G Log: {log_filename}")
        
        if not self.log_4g_dir:
            return result
        
        # 直接使用文件名查找（文件名是百分百准确的）
        log_file = self.log_4g_dir / log_filename
        
        if not log_file.exists():
            print(f"  文件不存在: {log_file}")
            return result
        
        # 只支持csv和xlsx格式
        if log_file.suffix.lower() not in ['.csv', '.xlsx']:
            print(f"  不支持的文件格式: {log_file.suffix}")
            return result
        
        print(f"  找到文件: {log_file}")
        
        try:
            # 判断文件格式
            if log_file.suffix.lower() == '.xlsx':
                # 处理XLSX格式
                wb = openpyxl.load_workbook(log_file, data_only=True)
                ws = wb.active
                
                # 读取表头
                headers = []
                header_row = 1
                for col_idx, cell in enumerate(ws[header_row], 1):
                    header_val = str(cell.value).strip().strip('"').strip("'") if cell.value else f'Col{col_idx}'
                    headers.append(header_val)
                
                # 查找字段索引
                def find_col_index(headers, field_key):
                    for idx, header in enumerate(headers):
                        if header.upper().strip() == field_key.upper().strip():
                            return idx + 1
                    return None
                
                rsrp_col = find_col_index(headers, 'RSRP')
                eci_col = find_col_index(headers, 'ECI')
                sinr_col = find_col_index(headers, 'SINR')
                lon_col = find_col_index(headers, 'LONGITUDE')
                lat_col = find_col_index(headers, 'LATITUDE')
                
                if not rsrp_col:
                    return result
                
                # 分两组：有经纬度的和无经纬度的，收集所有符合条件的行
                rows_with_loc = []  # [(rsrp, row_idx), ...]
                rows_without_loc = []  # [(rsrp, row_idx), ...]
                
                # 遍历所有数据行，只考虑CI、RSRP、SINR都不为空的行
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                    # 检查CI、RSRP、SINR是否都不为空
                    has_eci = eci_col and row[eci_col - 1].value is not None and str(row[eci_col - 1].value).strip() != ''
                    rsrp_cell = row[rsrp_col - 1]
                    has_rsrp = rsrp_cell.value is not None and str(rsrp_cell.value).strip() != ''
                    has_sinr = sinr_col and row[sinr_col - 1].value is not None and str(row[sinr_col - 1].value).strip() != ''
                    
                    # 必须CI、RSRP、SINR都不为空
                    if has_eci and has_rsrp and has_sinr:
                        try:
                            rsrp = float(rsrp_cell.value)
                            
                            # 检查是否有经纬度
                            has_lon = lon_col and row[lon_col - 1].value is not None and str(row[lon_col - 1].value).strip() != ''
                            has_lat = lat_col and row[lat_col - 1].value is not None and str(row[lat_col - 1].value).strip() != ''
                            has_location = has_lon and has_lat
                            
                            if has_location:
                                # 有经纬度的组
                                rows_with_loc.append((rsrp, row_idx))
                            else:
                                # 无经纬度的组
                                rows_without_loc.append((rsrp, row_idx))
                        except:
                            continue
                
                # 优先使用有经纬度的数据，取中间值
                selected_rows = rows_with_loc if rows_with_loc else rows_without_loc
                
                if selected_rows:
                    # 按 RSRP 值排序
                    selected_rows.sort(key=lambda x: x[0])
                    # 取中间值（中位数）
                    median_idx = len(selected_rows) // 2
                    median_row_idx = selected_rows[median_idx][1]
                    median_row = ws[median_row_idx]
                    def get_cell_value(headers, row, field_key):
                        col_idx = find_col_index(headers, field_key)
                        if col_idx:
                            cell = row[col_idx - 1]
                            val = str(cell.value).strip() if cell.value is not None else None
                            return val if val else None
                        return None
                    
                    result['经度'] = get_cell_value(headers, median_row, 'LONGITUDE')
                    result['纬度'] = get_cell_value(headers, median_row, 'LATITUDE')
                    result['ECI'] = get_cell_value(headers, median_row, 'ECI')
                    result['RSRP'] = get_cell_value(headers, median_row, 'RSRP')
                    result['SINR'] = get_cell_value(headers, median_row, 'SINR')
            else:
                # 处理CSV格式
                with open(log_file, 'r', encoding='utf-8-sig') as f:
                    # 先读取第一行，检查字段名
                    first_line = f.readline().strip()
                    f.seek(0)  # 重置文件指针
                    
                    reader = csv.DictReader(f)
                    # 获取实际的字段名（去除引号和空格）
                    fieldnames = [field.strip().strip('"').strip("'") for field in reader.fieldnames] if reader.fieldnames else []
                    
                    # 调试：打印字段名
                    print(f"4G Log字段名: {fieldnames}")
                    
                    # 查找字段索引
                    rsrp_col_idx = None
                    eci_col_idx = None
                    sinr_col_idx = None
                    lon_col_idx = None
                    lat_col_idx = None
                    for i, fieldname in enumerate(fieldnames):
                        fn_upper = fieldname.upper().strip()
                        if fn_upper == 'RSRP':
                            rsrp_col_idx = i
                        elif fn_upper == 'ECI':
                            eci_col_idx = i
                        elif fn_upper == 'SINR':
                            sinr_col_idx = i
                        elif fn_upper == 'LONGITUDE':
                            lon_col_idx = i
                        elif fn_upper == 'LATITUDE':
                            lat_col_idx = i
                    
                    # 分两组：有经纬度的和无经纬度的，收集所有有效的RSRP值
                    valid_rows_with_loc = []  # [(rsrp, row), ...]
                    valid_rows_without_loc = []  # [(rsrp, row), ...]
                    
                    for row in reader:
                        row_values = list(row.values())
                        
                        # 获取ECI值
                        eci_str = None
                        if eci_col_idx is not None and eci_col_idx < len(row_values):
                            eci_str = str(row_values[eci_col_idx]).strip().strip('"').strip("'")
                        if not eci_str:
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == 'ECI':
                                    eci_str = str(row[key]).strip().strip('"').strip("'")
                                    break
                        
                        # 获取RSRP值
                        rsrp_str = None
                        if rsrp_col_idx is not None and rsrp_col_idx < len(row_values):
                            rsrp_str = str(row_values[rsrp_col_idx]).strip().strip('"').strip("'")
                        if not rsrp_str:
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == 'RSRP':
                                    rsrp_str = str(row[key]).strip().strip('"').strip("'")
                                    break
                        
                        # 获取SINR值
                        sinr_str = None
                        if sinr_col_idx is not None and sinr_col_idx < len(row_values):
                            sinr_str = str(row_values[sinr_col_idx]).strip().strip('"').strip("'")
                        if not sinr_str:
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == 'SINR':
                                    sinr_str = str(row[key]).strip().strip('"').strip("'")
                                    break
                        
                        # 获取经纬度
                        lon_str = None
                        lat_str = None
                        if lon_col_idx is not None and lon_col_idx < len(row_values):
                            lon_str = str(row_values[lon_col_idx]).strip().strip('"').strip("'")
                        if lat_col_idx is not None and lat_col_idx < len(row_values):
                            lat_str = str(row_values[lat_col_idx]).strip().strip('"').strip("'")
                        
                        # 必须CI、RSRP、SINR都不为空
                        if eci_str and eci_str != '' and eci_str.lower() != 'none' and \
                           rsrp_str and rsrp_str != '' and rsrp_str.lower() != 'none' and \
                           sinr_str and sinr_str != '' and sinr_str.lower() != 'none':
                            try:
                                rsrp = float(rsrp_str)
                                
                                # 检查是否有有效经纬度
                                has_location = (lon_str and lon_str != '' and lon_str.lower() != 'none' and
                                               lat_str and lat_str != '' and lat_str.lower() != 'none')
                                
                                if has_location:
                                    valid_rows_with_loc.append((rsrp, row))
                                else:
                                    valid_rows_without_loc.append((rsrp, row))
                            except:
                                continue
                    
                    # 计算中位数并找到对应的行
                    median_row = None
                    if valid_rows_with_loc:
                        # 优先使用有经纬度的数据
                        valid_rows_with_loc.sort(key=lambda x: x[0])
                        median_idx = len(valid_rows_with_loc) // 2
                        median_row = valid_rows_with_loc[median_idx][1]
                    elif valid_rows_without_loc:
                        # 如果没有有经纬度的数据，使用无经纬度的数据
                        valid_rows_without_loc.sort(key=lambda x: x[0])
                        median_idx = len(valid_rows_without_loc) // 2
                        median_row = valid_rows_without_loc[median_idx][1]
                    
                    if median_row:
                        # 通过字段名索引匹配获取字段值
                        def get_field_value(row, fieldnames, field_key):
                            # 先通过字段名索引匹配
                            for i, fieldname in enumerate(fieldnames):
                                if fieldname.upper().strip() == field_key.upper().strip():
                                    row_values = list(row.values())
                                    if i < len(row_values):
                                        val = str(row_values[i]).strip().strip('"').strip("'")
                                        if val and val.lower() != 'none':
                                            return val
                            # 再尝试直接匹配
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == field_key.upper().strip():
                                    val = str(row[key]).strip().strip('"').strip("'")
                                    if val and val.lower() != 'none':
                                        return val
                            return None
                        
                        result['经度'] = get_field_value(median_row, fieldnames, 'LONGITUDE')
                        result['纬度'] = get_field_value(median_row, fieldnames, 'LATITUDE')
                        result['ECI'] = get_field_value(median_row, fieldnames, 'ECI')
                        result['RSRP'] = get_field_value(median_row, fieldnames, 'RSRP')
                        result['SINR'] = get_field_value(median_row, fieldnames, 'SINR')
                        
                        print(f"4G Log提取结果: 经度={result['经度']}, 纬度={result['纬度']}, RSRP={result['RSRP']}")
        except Exception as e:
            print(f"解析4G log失败 {log_file}: {e}")
        
        return result
    
    def parse_5g_log(self, log_filename: str) -> Dict:
        """解析5G log文件（支持CSV和XLSX格式）
        添加经纬度支持，优先选择有经纬度的数据
        """
        result = {
            '5G_经度': None,
            '5G_纬度': None,
            'NR-CI': None,
            'SS-RSRP': None,
            'SS-SINR': None
        }
        
        print(f"处理5G Log: {log_filename}")
        
        if not self.log_5g_dir:
            return result
        
        # 直接使用文件名查找（文件名是百分百准确的）
        log_file = self.log_5g_dir / log_filename
        
        if not log_file.exists():
            print(f"  文件不存在: {log_file}")
            return result
        
        # 只支持csv和xlsx格式
        if log_file.suffix.lower() not in ['.csv', '.xlsx']:
            print(f"  不支持的文件格式: {log_file.suffix}")
            return result
        
        print(f"  找到文件: {log_file}")
        
        try:
            # 判断文件格式
            if log_file.suffix.lower() == '.xlsx':
                # 处理XLSX格式
                wb = openpyxl.load_workbook(log_file, data_only=True)
                ws = wb.active
                
                # 读取表头
                headers = []
                header_row = 1
                for col_idx, cell in enumerate(ws[header_row], 1):
                    header_val = str(cell.value).strip().strip('"').strip("'") if cell.value else f'Col{col_idx}'
                    headers.append(header_val)
                
                # 查找字段索引
                def find_col_index(headers, field_key):
                    for idx, header in enumerate(headers):
                        if header.upper().strip() == field_key.upper().strip():
                            return idx + 1
                    return None
                
                rsrp_col = find_col_index(headers, 'SS-RSRP')
                ci_col = find_col_index(headers, 'NR-CI')
                sinr_col = find_col_index(headers, 'SS-SINR')
                lon_col = find_col_index(headers, 'LONGITUDE')
                lat_col = find_col_index(headers, 'LATITUDE')
                
                if not rsrp_col:
                    return result
                
                # 分两组：有经纬度的和无经纬度的，收集所有有效的RSRP值
                valid_rows_with_loc = []  # [(rsrp, row_idx), ...]
                valid_rows_without_loc = []  # [(rsrp, row_idx), ...]
                
                # 遍历所有数据行，只考虑CI、RSRP、SINR都不为空的行
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                    # 检查CI、RSRP、SINR是否都不为空
                    has_ci = ci_col and row[ci_col - 1].value is not None and str(row[ci_col - 1].value).strip() != ''
                    rsrp_cell = row[rsrp_col - 1]
                    has_rsrp = rsrp_cell.value is not None and str(rsrp_cell.value).strip() != ''
                    has_sinr = sinr_col and row[sinr_col - 1].value is not None and str(row[sinr_col - 1].value).strip() != ''
                    
                    # 必须CI、RSRP、SINR都不为空
                    if has_ci and has_rsrp and has_sinr:
                        try:
                            rsrp = float(rsrp_cell.value)
                            
                            # 检查是否有经纬度
                            has_lon = lon_col and row[lon_col - 1].value is not None and str(row[lon_col - 1].value).strip() != ''
                            has_lat = lat_col and row[lat_col - 1].value is not None and str(row[lat_col - 1].value).strip() != ''
                            has_location = has_lon and has_lat
                            
                            if has_location:
                                valid_rows_with_loc.append((rsrp, row_idx))
                            else:
                                valid_rows_without_loc.append((rsrp, row_idx))
                        except:
                            continue
                
                # 计算中位数并找到对应的行
                median_row_idx = None
                if valid_rows_with_loc:
                    # 优先使用有经纬度的数据
                    valid_rows_with_loc.sort(key=lambda x: x[0])
                    median_idx = len(valid_rows_with_loc) // 2
                    median_row_idx = valid_rows_with_loc[median_idx][1]
                elif valid_rows_without_loc:
                    # 如果没有有经纬度的数据，使用无经纬度的数据
                    valid_rows_without_loc.sort(key=lambda x: x[0])
                    median_idx = len(valid_rows_without_loc) // 2
                    median_row_idx = valid_rows_without_loc[median_idx][1]
                
                if median_row_idx:
                    median_row = ws[median_row_idx]
                    def get_cell_value(headers, row, field_key):
                        col_idx = find_col_index(headers, field_key)
                        if col_idx:
                            cell = row[col_idx - 1]
                            val = str(cell.value).strip() if cell.value is not None else None
                            return val if val else None
                        return None
                    
                    result['5G_经度'] = get_cell_value(headers, median_row, 'LONGITUDE')
                    result['5G_纬度'] = get_cell_value(headers, median_row, 'LATITUDE')
                    result['NR-CI'] = get_cell_value(headers, median_row, 'NR-CI')
                    result['SS-RSRP'] = get_cell_value(headers, median_row, 'SS-RSRP')
                    result['SS-SINR'] = get_cell_value(headers, median_row, 'SS-SINR')
            else:
                # 处理CSV格式
                with open(log_file, 'r', encoding='utf-8-sig') as f:
                    # 先读取第一行，检查字段名
                    first_line = f.readline().strip()
                    f.seek(0)  # 重置文件指针
                    
                    reader = csv.DictReader(f)
                    # 获取实际的字段名（去除引号和空格）
                    fieldnames = [field.strip().strip('"').strip("'") for field in reader.fieldnames] if reader.fieldnames else []
                    
                    # 调试：打印字段名
                    print(f"5G Log字段名: {fieldnames}")
                    
                    # 查找字段索引
                    rsrp_col_idx = None
                    ci_col_idx = None
                    sinr_col_idx = None
                    lon_col_idx = None
                    lat_col_idx = None
                    for i, fieldname in enumerate(fieldnames):
                        fn_upper = fieldname.upper().strip()
                        if fn_upper == 'SS-RSRP':
                            rsrp_col_idx = i
                        elif fn_upper == 'NR-CI':
                            ci_col_idx = i
                        elif fn_upper == 'SS-SINR':
                            sinr_col_idx = i
                        elif fn_upper == 'LONGITUDE':
                            lon_col_idx = i
                        elif fn_upper == 'LATITUDE':
                            lat_col_idx = i
                    
                    # 分两组：有经纬度的和无经纬度的，收集所有有效的RSRP值
                    valid_rows_with_loc = []  # [(rsrp, row), ...]
                    valid_rows_without_loc = []  # [(rsrp, row), ...]
                    
                    for row in reader:
                        row_values = list(row.values())
                        
                        # 获取NR-CI值
                        ci_str = None
                        if ci_col_idx is not None and ci_col_idx < len(row_values):
                            ci_str = str(row_values[ci_col_idx]).strip().strip('"').strip("'")
                        if not ci_str:
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == 'NR-CI':
                                    ci_str = str(row[key]).strip().strip('"').strip("'")
                                    break
                        
                        # 获取SS-RSRP值
                        rsrp_str = None
                        if rsrp_col_idx is not None and rsrp_col_idx < len(row_values):
                            rsrp_str = str(row_values[rsrp_col_idx]).strip().strip('"').strip("'")
                        if not rsrp_str:
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == 'SS-RSRP':
                                    rsrp_str = str(row[key]).strip().strip('"').strip("'")
                                    break
                        
                        # 获取SS-SINR值
                        sinr_str = None
                        if sinr_col_idx is not None and sinr_col_idx < len(row_values):
                            sinr_str = str(row_values[sinr_col_idx]).strip().strip('"').strip("'")
                        if not sinr_str:
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == 'SS-SINR':
                                    sinr_str = str(row[key]).strip().strip('"').strip("'")
                                    break
                        
                        # 获取经纬度
                        lon_str = None
                        lat_str = None
                        if lon_col_idx is not None and lon_col_idx < len(row_values):
                            lon_str = str(row_values[lon_col_idx]).strip().strip('"').strip("'")
                        if lat_col_idx is not None and lat_col_idx < len(row_values):
                            lat_str = str(row_values[lat_col_idx]).strip().strip('"').strip("'")
                        
                        # 必须CI、RSRP、SINR都不为空
                        if ci_str and ci_str != '' and ci_str.lower() != 'none' and \
                           rsrp_str and rsrp_str != '' and rsrp_str.lower() != 'none' and \
                           sinr_str and sinr_str != '' and sinr_str.lower() != 'none':
                            try:
                                rsrp = float(rsrp_str)
                                
                                # 检查是否有有效经纬度
                                has_location = (lon_str and lon_str != '' and lon_str.lower() != 'none' and
                                               lat_str and lat_str != '' and lat_str.lower() != 'none')
                                
                                if has_location:
                                    valid_rows_with_loc.append((rsrp, row))
                                else:
                                    valid_rows_without_loc.append((rsrp, row))
                            except:
                                continue
                    
                    # 计算中位数并找到对应的行
                    median_row = None
                    if valid_rows_with_loc:
                        # 优先使用有经纬度的数据
                        valid_rows_with_loc.sort(key=lambda x: x[0])
                        median_idx = len(valid_rows_with_loc) // 2
                        median_row = valid_rows_with_loc[median_idx][1]
                    elif valid_rows_without_loc:
                        # 如果没有有经纬度的数据，使用无经纬度的数据
                        valid_rows_without_loc.sort(key=lambda x: x[0])
                        median_idx = len(valid_rows_without_loc) // 2
                        median_row = valid_rows_without_loc[median_idx][1]
                    
                    if median_row:
                        # 通过字段名索引匹配获取字段值
                        def get_field_value(row, fieldnames, field_key):
                            # 先通过字段名索引匹配
                            for i, fieldname in enumerate(fieldnames):
                                if fieldname.upper().strip() == field_key.upper().strip():
                                    row_values = list(row.values())
                                    if i < len(row_values):
                                        val = str(row_values[i]).strip().strip('"').strip("'")
                                        if val and val.lower() != 'none':
                                            return val
                            # 再尝试直接匹配
                            for key in row.keys():
                                if key.strip().strip('"').strip("'").upper() == field_key.upper().strip():
                                    val = str(row[key]).strip().strip('"').strip("'")
                                    if val and val.lower() != 'none':
                                        return val
                            return None
                        
                        result['5G_经度'] = get_field_value(median_row, fieldnames, 'LONGITUDE')
                        result['5G_纬度'] = get_field_value(median_row, fieldnames, 'LATITUDE')
                        result['NR-CI'] = get_field_value(median_row, fieldnames, 'NR-CI')
                        result['SS-RSRP'] = get_field_value(median_row, fieldnames, 'SS-RSRP')
                        result['SS-SINR'] = get_field_value(median_row, fieldnames, 'SS-SINR')
                        
                        print(f"5G Log提取结果: 经度={result['5G_经度']}, 纬度={result['5G_纬度']}, SS-RSRP={result['SS-RSRP']}")
        except Exception as e:
            print(f"解析5G log失败 {log_file}: {e}")
        
        return result
    
    def save_results(self, results: List[Dict], output_path: Path):
        """保存结果到CSV"""
        if not results:
            return
        
        fieldnames = ['工单号', '经度', '纬度', '上传速率Mbps', '下载速率Mbps', 
                     'ECI', 'RSRP', 'SINR', 'NR-CI', 'SS-RSRP', 'SS-SINR']
        
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for result in results:
                writer.writerow({k: result.get(k, '') for k in fieldnames})
